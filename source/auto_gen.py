import openpyxl
import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32
import os

# xls에서 xlsx로 변환.
filename = input("파일명을 입력하세요: ")
path = os.getcwd()

fname = f"{path}\\{filename}.xls"
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)    
wb.Close()                              
excel.Application.Quit()

# 생성된 xlsx 파일의 정보를 수집 및 오름차순으로 분류. 
temp = pd.read_excel(f"{filename}.xlsx")
df = temp.loc[: ,["대리점", "모델명", "수취인", "주문번호"]][:-2]
df['주문번호'] = pd.to_numeric(df["주문번호"], errors = "ignore")
count = df.value_counts()

duplicate_count_df = df.groupby(df.columns.tolist(), as_index=False).size().drop_duplicates(subset=list(df.columns))
final = duplicate_count_df

final.sort_values(by="주문번호", ascending = False, inplace = True)
final.set_index("대리점", inplace = True)

# 추출된 데이터를 xlsx로 같은 경로에 생성.
address = f"{path}\\{filename}.xlsx"
final.to_excel(excel_writer = address, sheet_name="판매 통계")

print("엑셀 데이터 파일 생성이 완료 되었습니다.")

# 엑셀 마지막 자리 지정.
end = len(final) + 1

# 생성된 데이터 불러오기.
bxm = load_workbook(f'{filename}.xlsx')
bxm_ws = bxm.active

# 필요한 정보 수집 및 할당.
order = bxm_ws[f"D{end}"].value
temp = str(order)
gen_date = f"20" + temp[:2] + f"-" + temp[2:4] + f"-" + temp[4:6]
model = bxm_ws[f"B{end}"].value
receiver = bxm_ws[f"A{end}"].value
receiver = receiver[7:]
lot = bxm_ws[f"E{end}"].value
lot_num = 1

wb = load_workbook('공정검사 format.xlsx')
ws = wb.active
ws.title = str(order)

eb = load_workbook('출하검사 format.xlsx')
es = eb.active
es.title = str(order)

# 새로 생성한 데이터를 자동으로 입력.
ws["E4"] = model            
ws["E5"] = lot                 
ws["P4"] = gen_date
ws["P5"] = lot_num
ws["P6"] = receiver


es["E5"] = model            
es["E6"] = lot            
es["R5"] = gen_date
es["R6"] = gen_date
es["E7"] = receiver

if lot == 1:
    ws["I10"] = "n=1, c=0"
    ws["I11"] = "n=1, c=0"
    ws["I14"] = "n=1, c=0"
    ws["I17"] = "n=1, c=0"
    ws["I18"] = "n=1, c=0"
    ws["I19"] = "n=1, c=0"
    ws["I20"] = "n=1, c=0"
    ws["I21"] = "n=1, c=0"
    ws["I23"] = "n=1, c=0"
    ws["I29"] = "n=1, c=0"
    ws["I30"] = "n=1, c=0"
    
    es["I12"] = "n=1, c=0"
    es["I13"] = "n=1, c=0"
    es["I14"] = "n=1, c=0"
    es["I15"] = "n=1, c=0"
    es["I16"] = "n=1, c=0"
    es["I17"] = "n=1, c=0"
    es["I18"] = "n=1, c=0"
    es["I19"] = "n=1, c=0"
    es["I20"] = "n=1, c=0"
    es["I21"] = "n=1, c=0"
    es["I22"] = "n=1, c=0"
    es["I23"] = "n=1, c=0"
    es["I24"] = "n=1, c=0"
    es["I27"] = "n=1, c=0"
    es["I28"] = "n=1, c=0"
    
elif lot == 2:
    ws["I10"] = "n=2, c=0"
    ws["I11"] = "n=2, c=0"
    ws["I14"] = "n=2, c=0"
    ws["I17"] = "n=2, c=0"
    ws["I18"] = "n=2, c=0"
    ws["I19"] = "n=2, c=0"
    ws["I20"] = "n=2, c=0"
    ws["I21"] = "n=2, c=0"
    ws["I23"] = "n=2, c=0"
    ws["I29"] = "n=2, c=0"
    ws["I30"] = "n=2, c=0"
    
    es["I12"] = "n=2, c=0"
    es["I13"] = "n=2, c=0"
    es["I14"] = "n=2, c=0"
    es["I15"] = "n=2, c=0"
    es["I16"] = "n=2, c=0"
    es["I17"] = "n=2, c=0"
    es["I18"] = "n=2, c=0"
    es["I19"] = "n=2, c=0"
    es["I20"] = "n=2, c=0"
    es["I21"] = "n=2, c=0"
    es["I22"] = "n=2, c=0"
    es["I23"] = "n=2, c=0"
    es["I24"] = "n=2, c=0"
    es["I27"] = "n=2, c=0"
    es["I28"] = "n=2, c=0"
    
elif lot > 2:
    ws["I10"] = "n=3, c=0"
    ws["I11"] = "n=3, c=0"
    ws["I14"] = "n=3, c=0"
    ws["I17"] = "n=3, c=0"
    ws["I18"] = "n=3, c=0"
    ws["I19"] = "n=3, c=0"
    ws["I20"] = "n=3, c=0"
    ws["I21"] = "n=3, c=0"
    ws["I23"] = "n=3, c=0"
    ws["I29"] = "n=3, c=0"
    ws["I30"] = "n=3, c=0"
    
    es["I12"] = "n=3, c=0"
    es["I13"] = "n=3, c=0"
    es["I14"] = "n=3, c=0"
    es["I15"] = "n=3, c=0"
    es["I16"] = "n=3, c=0"
    es["I17"] = "n=3, c=0"
    es["I18"] = "n=3, c=0"
    es["I19"] = "n=3, c=0"
    es["I20"] = "n=3, c=0"
    es["I21"] = "n=3, c=0"
    es["I22"] = "n=3, c=0"
    es["I23"] = "n=3, c=0"
    es["I24"] = "n=3, c=0"
    es["I27"] = "n=3, c=0"
    es["I28"] = "n=3, c=0"
    

# 위 내용을 새로운 시트를 생성할 때 마다 반복.
for i in range(end-1, 1, -1):
        
    target1 = wb.copy_worksheet(ws)
    target2 = eb.copy_worksheet(es)

    order = bxm_ws[f"D{i}"].value
    temp = str(order)
    gen_date = f"20" + temp[:2] + f"-" + temp[2:4] + f"-" + temp[4:6]
    model = bxm_ws[f"B{i}"].value
    receiver = bxm_ws[f"A{i}"].value
    receiver = receiver[7:]
    lot_num += 1
    lot = bxm_ws[f"E{i}"].value

    target1.title = str(order)
    target2.title = str(order)

    
    target1.title = str(order)
    target1["E4"] = model            
    target1["E5"] = lot                 
    target1["P4"] = gen_date
    target1["P5"] = lot_num
    target1["P6"] = receiver

    target2.title = str(order)
    target2["E5"] = model           
    target2["E6"] = lot                 
    target2["R5"] = gen_date
    target2["R6"] = gen_date
    target2["E7"] = receiver

    if lot == 1:
        target1["I10"] = "n=1, c=0"
        target1["I11"] = "n=1, c=0"
        target1["I14"] = "n=1, c=0"
        target1["I17"] = "n=1, c=0"
        target1["I18"] = "n=1, c=0"
        target1["I19"] = "n=1, c=0"
        target1["I20"] = "n=1, c=0"
        target1["I21"] = "n=1, c=0"
        target1["I23"] = "n=1, c=0"
        target1["I29"] = "n=1, c=0"
        target1["I30"] = "n=1, c=0"

        target2["I12"] = "n=1, c=0"
        target2["I13"] = "n=1, c=0"
        target2["I14"] = "n=1, c=0"
        target2["I15"] = "n=1, c=0"
        target2["I16"] = "n=1, c=0"
        target2["I17"] = "n=1, c=0"
        target2["I18"] = "n=1, c=0"
        target2["I19"] = "n=1, c=0"
        target2["I20"] = "n=1, c=0"
        target2["I21"] = "n=1, c=0"
        target2["I22"] = "n=1, c=0"
        target2["I23"] = "n=1, c=0"
        target2["I24"] = "n=1, c=0"
        target2["I27"] = "n=1, c=0"
        target2["I28"] = "n=1, c=0"

    elif lot == 2:
        target1["I10"] = "n=2, c=0"
        target1["I11"] = "n=2, c=0"
        target1["I14"] = "n=2, c=0"
        target1["I17"] = "n=2, c=0"
        target1["I18"] = "n=2, c=0"
        target1["I19"] = "n=2, c=0"
        target1["I20"] = "n=2, c=0"
        target1["I21"] = "n=2, c=0"
        target1["I23"] = "n=2, c=0"
        target1["I29"] = "n=2, c=0"
        target1["I30"] = "n=2, c=0"

        target2["I12"] = "n=2, c=0"
        target2["I13"] = "n=2, c=0"
        target2["I14"] = "n=2, c=0"
        target2["I15"] = "n=2, c=0"
        target2["I16"] = "n=2, c=0"
        target2["I17"] = "n=2, c=0"
        target2["I18"] = "n=2, c=0"
        target2["I19"] = "n=2, c=0"
        target2["I20"] = "n=2, c=0"
        target2["I21"] = "n=2, c=0"
        target2["I22"] = "n=2, c=0"
        target2["I23"] = "n=2, c=0"
        target2["I24"] = "n=2, c=0"
        target2["I27"] = "n=2, c=0"
        target2["I28"] = "n=2, c=0"
        
    elif lot > 2:
        target1["I10"] = "n=3, c=0"
        target1["I11"] = "n=3, c=0"
        target1["I14"] = "n=3, c=0"
        target1["I17"] = "n=3, c=0"
        target1["I18"] = "n=3, c=0"
        target1["I19"] = "n=3, c=0"
        target1["I20"] = "n=3, c=0"
        target1["I21"] = "n=3, c=0"
        target1["I23"] = "n=3, c=0"
        target1["I29"] = "n=3, c=0"
        target1["I30"] = "n=3, c=0"

        target2["I12"] = "n=3, c=0"
        target2["I13"] = "n=3, c=0"
        target2["I14"] = "n=3, c=0"
        target2["I15"] = "n=3, c=0"
        target2["I16"] = "n=3, c=0"
        target2["I17"] = "n=3, c=0"
        target2["I18"] = "n=3, c=0"
        target2["I19"] = "n=3, c=0"
        target2["I20"] = "n=3, c=0"
        target2["I21"] = "n=3, c=0"
        target2["I22"] = "n=3, c=0"
        target2["I23"] = "n=3, c=0"
        target2["I24"] = "n=3, c=0"
        target2["I27"] = "n=3, c=0"
        target2["I28"] = "n=3, c=0"


# 엑셀 이름 설정.
excel_name1 = f"{path}\\result\\공정검사성적서(데스크탑) {gen_date[:7]}.xlsx"
excel_name2 = f"{path}\\result\\출하검사성적서(데스크탑) {gen_date[:7]}.xlsx"


# 생성된 모든 데이터 저장 및 엑셀 파일 생성.
wb.save(excel_name1)
print("공정검사성적서 생성이 완료 되었습니다.")

eb.save(excel_name2)
print("출하검사성적서 생성이 완료 되었습니다.")

# 작업 완료 출력.
print()
print("모든 작업이 완료 되었습니다.")
